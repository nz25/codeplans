from win32com import client
from enum import IntEnum
from collections import namedtuple


RawVariableInfo = namedtuple('VariableInfo', 'name label data_type categories')
MasterVariableInfo = namedtuple('VariableInfo', 'name data_type v_new v_dropped c_new c_dropped v_label c_label')

class DataTypeConstants(IntEnum):
    mtNone = 0
    mtLong = 1
    mtText = 2
    mtCategorical = 3
    mtObject = 4
    mtDate = 5
    mtDouble = 6
    mtBoolean = 7

class openConstants(IntEnum):
     oREAD      = 1
     oREADWRITE = 2
     oNOSAVE    = 3

def get_mdd_data(mdd_path):

    mdd = client.Dispatch('MDM.Document')
    mdd.Open(mdd_path, mode=openConstants.oREAD)

    variables = {}
    for v in mdd.Variables:
        name = v.fullname
        label = str(v.fulllabel)
        data_type=str(DataTypeConstants(v.datatype)).split('.')[1]
        categories = {}
        for c in v.Categories:
            categories[c.Name] = c.Label
        variables[name] = RawVariableInfo(name, label, data_type, categories)
    
    mdd.Close()
    return variables

def create_excel_comparison(old_mdd, new_mdd, xl_output):
    old_wave_variables = get_mdd_data(old_mdd)
    new_wave_variables = get_mdd_data(new_mdd)

    # filling master variables
    master_variables = []

    # 1. new variables
    for name, variable in new_wave_variables.items():
        if name not in old_wave_variables:
            master_variables.append(
                MasterVariableInfo(
                    name, variable.data_type,
                    v_new=True,
                    v_dropped=False,
                    c_new=False,
                    c_dropped=False,
                    v_label=False,
                    c_label=False)
            )

    # 2. dropped variables
    for name, variable in old_wave_variables.items():
        if name not in new_wave_variables:
            master_variables.append(
                MasterVariableInfo(
                    name, variable.data_type,
                    v_new=False,
                    v_dropped=True,
                    c_new=False,
                    c_dropped=False,
                    v_label=False,
                    c_label=False)
            )

    # 3. changed variables

    for new_variable in new_wave_variables.values():
        old_variable = old_wave_variables.get(new_variable.name)

        if old_variable:

            # 3.1 check variable labels
            v_label = old_variable.label != new_variable.label

            # 3.2 check categories
            for new_name, new_label in new_variable.categories.items():
                old_label = old_variable.categories.get(new_name)
                if old_label and new_label != old_label:
                    c_label = True
                    break
            else:
                c_label = False

            # 3.3 new categories
            c_new = bool([c for c in new_variable.categories if c not in old_variable.categories])
            c_dropped = bool([c for c in old_variable.categories if c not in new_variable.categories])

            if v_label or c_label or c_new or c_dropped:
                master_variables.append(
                    MasterVariableInfo(
                    name=new_variable.name,
                    data_type=new_variable.data_type,
                    v_new=False,
                    v_dropped=False,
                    c_new=c_new,
                    c_dropped=c_dropped,
                    v_label=v_label,
                    c_label=c_label)
                )

    # export in excel

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Side, PatternFill
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import colors
    from openpyxl.styles import Font, Color

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "overview"

    # report header (row 1-2)

    ws1.append(['New', new_mdd])
    ws1.append(['Old', old_mdd])
    ws1.append([])

    if master_variables:

        # variable header (row 4-5)
        ws1.append(['name', 'data_type', 'variable', '', 'categories', '', 'label(s) changed', '', 'Remarks'])
        ws1.append(['', '', 'new', 'dropped', 'new', 'dropped', 'variable', 'categories', ''])

        for v in master_variables:
            booleans = map(lambda x: 'x' if x else '', v[2:])
            # booleans = ['x' if x else '' for x in v[2:]]
            ws1.append((v.name, v.data_type, *booleans))

        #formatting

        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 20

        # alignment - header
        for row in range(4, 6):
            for col in range(1, 12):
                ws1.cell(column=col, row=row).alignment = Alignment(horizontal='center', vertical='center')

        ws1.merge_cells('A4:A5')
        ws1.merge_cells('B4:B5')
        ws1.merge_cells('C4:D4')
        ws1.merge_cells('E4:F4')
        ws1.merge_cells('G4:H4')
        ws1.merge_cells('I4:I5')

        # borders
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
        # create fill
        colour_fill = PatternFill(start_color='FFFF7F',
                                end_color='FFFF7F',
                                fill_type='solid')  
    
        for row in range(4, len(master_variables) + 6):
            for col in range(1, len(master_variables[0]) + 2):
                ws1.cell(row=row, column=col).border = thin_border
            # alignment - table        
            for col in range(3, len(master_variables[0]) + 2):
                ws1.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')

        # freeze header
        ws1.freeze_panes = ws1.cell(row=6, column=1)

        # conditional formatting
        ws1.conditional_formatting.add(f'B6:B{len(master_variables) + 5}',
                        FormulaRule(formula=['B6="mtText"'], stopIfTrue=True, fill=colour_fill))

        # auto filter
        ws1.auto_filter.ref = f'A5:I{len(master_variables) + 4}'

    else:
        ws1.append(['... no changes in variables'])
        ws1.append(['... check routing, use toolbox.bat'])

        a1 = ws1['A4']
        a2 = ws1['A5']
        ft = Font(color=colors.RED)
        a1.font=ft
        a2.font=ft

    wb.save(filename = xl_output)